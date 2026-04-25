"""Tests for the data scout.

Use fake adapters returning canned EventRecord lists; verify the three
output files are written and well-formed. No live network.
"""

import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from lawtracker.scout import run
from lawtracker.sources import EventRecord, PollResult, SourceAdapter


class _OkAdapter(SourceAdapter):
    source_id = "fake_ok"
    kind = "event_list"
    url = "https://example.test/ok"

    def poll(self, *, client=None) -> PollResult:  # type: ignore[override]
        return PollResult(
            status="ok",
            events=[
                EventRecord(
                    dedup_key="https://example.test/case/1",
                    source_id="fake_ok",
                    event_date=date(2026, 3, 15),
                    title="Test Case A",
                    primary_actor="Acme Corp",
                    summary="Quick summary",
                    url="https://example.test/case/1",
                    country="US",
                    metadata={"industry": "pharma", "case_number": "1:26-cr-00001"},
                ),
                EventRecord(
                    dedup_key="https://example.test/case/2",
                    source_id="fake_ok",
                    event_date=date(2025, 11, 10),
                    title="Test Case B",
                    primary_actor="Beta Inc",
                    summary=None,
                    url="https://example.test/case/2",
                    country="AU",
                    metadata={"industry": "extractive"},
                ),
            ],
        )

    def parse(self, html: str, client=None) -> list[EventRecord]:
        return []


class _OtherOkAdapter(SourceAdapter):
    source_id = "fake_blog"
    kind = "event_list"
    url = "https://example.test/blog"

    def poll(self, *, client=None) -> PollResult:  # type: ignore[override]
        return PollResult(
            status="ok",
            events=[
                EventRecord(
                    dedup_key="https://blog.test/post/1",
                    source_id="fake_blog",
                    event_date=date(2026, 4, 20),
                    title="Post Title",
                    primary_actor=None,
                    summary="Post summary",
                    url="https://blog.test/post/1",
                    country=None,
                    metadata={"tags": "FCPA, anticorruption"},
                ),
            ],
        )

    def parse(self, html: str, client=None) -> list[EventRecord]:
        return []


class _FailingAdapter(SourceAdapter):
    source_id = "fake_fail"
    kind = "event_list"
    url = "https://example.test/fail"

    def poll(self, *, client=None) -> PollResult:  # type: ignore[override]
        return PollResult(status="permanent_failure", error="simulated failure")

    def parse(self, html: str, client=None) -> list[EventRecord]:
        return []


def test_scout_writes_three_files(tmp_path: Path):
    report = run([_OkAdapter, _OtherOkAdapter], tmp_path)
    assert (tmp_path / "events.xlsx").exists()
    assert (tmp_path / "events.jsonl").exists()
    assert (tmp_path / "summary.txt").exists()
    assert report["events_collected"] == 3


def test_xlsx_has_universal_columns_and_metadata_union(tmp_path: Path):
    run([_OkAdapter, _OtherOkAdapter], tmp_path)

    wb = load_workbook(tmp_path / "events.xlsx")
    ws = wb.active
    headers = [c.value for c in ws[1]]

    universal = [
        "event_date",
        "source_id",
        "country",
        "title",
        "primary_actor",
        "summary",
        "url",
        "dedup_key",
    ]
    assert headers[:8] == universal

    metadata_cols = headers[8:]
    assert metadata_cols == sorted(metadata_cols)
    assert "industry" in metadata_cols
    assert "case_number" in metadata_cols
    assert "tags" in metadata_cols
    assert ws.freeze_panes == "A2"


class _PodcastInDisguiseAdapter(SourceAdapter):
    """An entry whose title doesn't reveal its podcast nature, but whose
    LLM-generated summary will. Models the M&C 'EMBARGOED!: South of the
    Border' case Tom flagged 2026-04-25."""

    source_id = "fake_disguised"
    kind = "event_list"
    url = "https://example.test/embargoed-feed"

    def poll(self, *, client=None) -> PollResult:  # type: ignore[override]
        return PollResult(
            status="ok",
            events=[
                EventRecord(
                    dedup_key="https://example.test/embargoed/south-border",
                    source_id="fake_disguised",
                    event_date=date(2025, 11, 14),
                    title="EMBARGOED!: South of the Border",
                    primary_actor=None,
                    # Adapter sees no summary at parse time — only the LLM
                    # later reads the article and writes a summary that
                    # mentions "podcast".
                    summary=(
                        "Miller & Chevalier's EMBARGOED! podcast discusses "
                        "sanctions developments in Latin America."
                    ),
                    url="https://example.test/embargoed/south-border",
                    country="US",
                ),
                EventRecord(
                    dedup_key="https://example.test/keep",
                    source_id="fake_disguised",
                    event_date=date(2025, 11, 14),
                    title="DOJ resolves $50M FCPA case with Acme Corp",
                    primary_actor="Acme Corp",
                    summary="Substantive enforcement news with no event-noise markers.",
                    url="https://example.test/keep",
                    country="US",
                ),
            ],
        )

    def parse(self, html: str, client=None) -> list[EventRecord]:
        return []


def test_post_enrichment_filter_drops_llm_revealed_podcasts(tmp_path: Path, monkeypatch):
    """Tom 2026-04-25: M&C 'EMBARGOED!' entries slipped through because
    their titles don't mention 'podcast' — the LLM revealed the nature in
    the summary. scout.run must re-filter after enrichment.

    `off` mode keeps the (already-podcast-mentioning) summary unchanged so
    the post-filter pass sees what an anthropic-mode summary would
    typically reveal: the word `podcast` somewhere in the body.
    """
    monkeypatch.setenv("LAWTRACKER_LLM_MODE", "off")
    report = run([_PodcastInDisguiseAdapter], tmp_path)
    assert report["events_collected"] == 1, (
        "podcast-summary entry must be dropped post-enrichment; only the "
        "substantive Acme Corp entry should remain"
    )

    parsed = [
        json.loads(line)
        for line in (tmp_path / "events.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    titles = [e["title"] for e in parsed]
    assert any("Acme Corp" in t for t in titles)
    assert not any("EMBARGOED" in t for t in titles)


def test_xlsx_title_column_has_hyperlink_to_url(tmp_path: Path):
    """Ellen 2026-04-25: clicking the title in Excel should open the source URL."""
    run([_OkAdapter, _OtherOkAdapter], tmp_path)

    wb = load_workbook(tmp_path / "events.xlsx")
    ws = wb.active
    headers = [c.value for c in ws[1]]
    title_idx = headers.index("title") + 1
    url_idx = headers.index("url") + 1

    for row_num in range(2, ws.max_row + 1):
        title_cell = ws.cell(row=row_num, column=title_idx)
        url_value = ws.cell(row=row_num, column=url_idx).value
        assert title_cell.hyperlink is not None, "title cell should be hyperlinked"
        assert title_cell.hyperlink.target == url_value


def test_jsonl_one_line_per_event(tmp_path: Path):
    run([_OkAdapter, _OtherOkAdapter], tmp_path)
    lines = (tmp_path / "events.jsonl").read_text(encoding="utf-8").strip().splitlines()
    assert len(lines) == 3
    parsed = [json.loads(line) for line in lines]
    assert all("dedup_key" in r for r in parsed)


def test_summary_includes_per_source_status_and_failure_error(tmp_path: Path):
    run([_OkAdapter, _OtherOkAdapter, _FailingAdapter], tmp_path)
    summary = (tmp_path / "summary.txt").read_text(encoding="utf-8")
    assert "fake_ok" in summary
    assert "fake_blog" in summary
    assert "fake_fail" in summary
    assert "permanent_failure" in summary
    assert "simulated failure" in summary


def test_summary_includes_country_and_industry_breakdowns(tmp_path: Path):
    run([_OkAdapter, _OtherOkAdapter], tmp_path)
    summary = (tmp_path / "summary.txt").read_text(encoding="utf-8")
    assert "Events per country" in summary
    assert "US" in summary
    assert "AU" in summary
    assert "(none)" in summary
    assert "Top 20 industries" in summary
    assert "pharma" in summary
    assert "extractive" in summary


def test_empty_run_still_writes_files(tmp_path: Path):
    report = run([_FailingAdapter], tmp_path)
    assert (tmp_path / "events.xlsx").exists()
    assert (tmp_path / "events.jsonl").exists()
    assert (tmp_path / "summary.txt").exists()
    assert report["events_collected"] == 0


def test_source_filter_restricts_to_one_adapter(tmp_path: Path):
    report = run(
        [_OkAdapter, _OtherOkAdapter, _FailingAdapter],
        tmp_path,
        source_filter="fake_ok",
    )
    assert report["events_collected"] == 2
    assert len(report["poll_log"]) == 1
    assert report["poll_log"][0]["source_id"] == "fake_ok"
