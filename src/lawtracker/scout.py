"""Data scout — one-shot adapter run with Excel / JSONL / summary outputs.

Runs every adapter in `PILOT_ADAPTERS` once, collects EventRecords, writes
them to disk for human review. No DB, no scheduling, no state between runs.

Three invocation styles:
    lawtracker scout
    python -m lawtracker scout
    python -m lawtracker.scout
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from lawtracker.sources import EventRecord, SourceAdapter
from lawtracker.sources._filters import matches_event_noise
from lawtracker.sources.afp_foreign_bribery import AfpForeignBriberyAdapter
from lawtracker.sources.consejo_transparencia_cl import ConsejoTransparenciaClAdapter
from lawtracker.sources.doj_fcpa_actions import DojFcpaActionsAdapter
from lawtracker.sources.fiscalia_chile import FiscaliaChileAdapter
from lawtracker.sources.foley_llp import FoleyLlpAdapter
from lawtracker.sources.gibson_dunn import GibsonDunnAdapter
from lawtracker.sources.global_anticorruption_blog import GlobalAnticorruptionBlogAdapter
from lawtracker.sources.harvard_corpgov_fcpa import HarvardCorpGovFcpaAdapter
from lawtracker.sources.miller_chevalier import MillerChevalierFcpaAdapter
from lawtracker.sources.sec_fcpa_cases import SecFcpaCasesAdapter
from lawtracker.sources.volkov_law import VolkovLawAdapter

PILOT_ADAPTERS: list[type[SourceAdapter]] = [
    DojFcpaActionsAdapter,
    SecFcpaCasesAdapter,
    AfpForeignBriberyAdapter,
    FiscaliaChileAdapter,
    ConsejoTransparenciaClAdapter,
    VolkovLawAdapter,
    GibsonDunnAdapter,
    GlobalAnticorruptionBlogAdapter,
    MillerChevalierFcpaAdapter,
    FoleyLlpAdapter,
    HarvardCorpGovFcpaAdapter,
    # Sources flagged blocked / JS-rendered / no-RSS: FCPA Blog (CDN 401),
    # SEC FCPA cases (CDN 403), OECD WGB (CDN 403), AUSTRAC / NACC / CDPP
    # (timeout — likely AU geo-block), ASIC (JS-rendered, no inline data),
    # WilmerHale / Sidley / Skadden / Debevoise / Latham / Cleary / Hogan
    # Lovells / Freshfields / Herbert Smith Freehills / DLA Piper /
    # Foley LLP / Foley Hoag / Covington / Ropes & Gray (no exposed RSS or
    # CDN 403/404). See design/sources.md and design/data-scout.md
    # "Findings" sections for the full audit.
]

DEFAULT_OUTPUT_DIR = Path("data/scout")

UNIVERSAL_COLUMNS: tuple[str, ...] = (
    "event_date",
    "source_id",
    "country",
    "title",
    "primary_actor",
    "summary",
    "url",
    "dedup_key",
)


def run(
    adapter_classes: list[type[SourceAdapter]],
    output_dir: Path,
    *,
    source_filter: str | None = None,
) -> dict[str, Any]:
    """Run the scout end-to-end. Returns a small report for the CLI."""
    output_dir.mkdir(parents=True, exist_ok=True)

    poll_log: list[dict[str, Any]] = []
    all_events: list[EventRecord] = []

    for cls in adapter_classes:
        if source_filter is not None and cls.source_id != source_filter:
            continue
        adapter = cls()
        result = adapter.poll()
        poll_log.append(
            {
                "source_id": cls.source_id,
                "status": result.status,
                "event_count": len(result.events),
                "error": result.error,
            }
        )
        if result.status == "ok":
            all_events.extend(result.events)

    all_events = _enrich_summaries(all_events, output_dir)

    # Re-apply event-noise filter after summary enrichment. Some entries
    # (e.g. M&C's "EMBARGOED!: South of the Border" podcast publications)
    # have innocuous titles that only reveal their podcast / webinar
    # nature once the LLM reads the article. The first filter pass at
    # parse time can't see that; this catches it.
    all_events = [
        e
        for e in all_events
        if not matches_event_noise(e.title, e.summary, e.url)
    ]

    _write_xlsx(all_events, output_dir / "events.xlsx")
    _write_jsonl(all_events, output_dir / "events.jsonl")
    _write_summary(all_events, poll_log, output_dir / "summary.txt")
    _write_analysis(all_events, output_dir / "analysis.md")

    return {
        "events_collected": len(all_events),
        "poll_log": poll_log,
        "output_dir": output_dir,
    }


def _write_analysis(events: list[EventRecord], path: Path) -> None:
    from lawtracker.analysis import build_analysis

    path.write_text(build_analysis(events), encoding="utf-8")


def _enrich_summaries(events: list[EventRecord], output_dir: Path) -> list[EventRecord]:
    """Per-event summary enrichment with disk-backed cache.

    Cache lives under `output_dir/.cache/summaries.json` so it shares the
    same gitignored data tree as everything else the scout produces.
    """
    from lawtracker.article_summary import enrich_summaries
    from lawtracker.llm_cache import JsonCache

    cache = JsonCache(output_dir / ".cache" / "summaries.json")
    return enrich_summaries(events, cache=cache)


def _write_xlsx(events: list[EventRecord], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("events")
    else:
        ws.title = "events"

    metadata_keys = sorted({k for e in events for k in e.metadata})
    columns = list(UNIVERSAL_COLUMNS) + metadata_keys

    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for event in events:
        row: list[Any] = []
        for col in columns:
            if col in UNIVERSAL_COLUMNS:
                value = getattr(event, col, None)
                if isinstance(value, date):
                    value = value.isoformat()
            else:
                value = event.metadata.get(col)
            row.append(value)
        ws.append(row)

    # Hyperlink the title column to each event's URL so Ellen can click through
    # straight from the spreadsheet (Ellen's review 2026-04-25).
    title_idx = columns.index("title") + 1
    url_idx = columns.index("url") + 1
    link_font = Font(color="0563C1", underline="single")
    for row_num in range(2, ws.max_row + 1):
        url_cell = ws.cell(row=row_num, column=url_idx)
        title_cell = ws.cell(row=row_num, column=title_idx)
        if url_cell.value:
            title_cell.hyperlink = url_cell.value
            title_cell.font = link_font

    for i, _ in enumerate(columns, start=1):
        letter = get_column_letter(i)
        max_len = max(
            (
                len(str(ws.cell(row=r, column=i).value or ""))
                for r in range(1, ws.max_row + 1)
            ),
            default=0,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

    wb.save(path)


def _write_jsonl(events: list[EventRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        for event in events:
            f.write(event.model_dump_json() + "\n")


def _write_summary(
    events: list[EventRecord],
    poll_log: list[dict[str, Any]],
    path: Path,
) -> None:
    lines: list[str] = []
    lines.append(f"Scout run at {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"Total events collected: {len(events)}")
    lines.append("")

    lines.extend(_section_per_source_totals(events, poll_log))
    lines.extend(_section_events_per_month(events))
    lines.extend(
        _section_count_breakdown(
            "Events per country",
            events,
            lambda e: e.country or "(none)",
        )
    )
    lines.extend(
        _section_count_breakdown(
            "Top 20 industries",
            events,
            lambda e: e.metadata.get("industry"),
            limit=20,
        )
    )
    lines.extend(
        _section_count_breakdown(
            "Top 20 primary_actors",
            events,
            lambda e: e.primary_actor,
            limit=20,
        )
    )

    path.write_text("\n".join(lines), encoding="utf-8")


def _section_per_source_totals(
    events: list[EventRecord], poll_log: list[dict[str, Any]]
) -> list[str]:
    out = ["== Per-source totals =="]
    for entry in poll_log:
        sid = str(entry["source_id"])
        status = str(entry["status"])
        count = int(entry["event_count"])
        last_date = max(
            (e.event_date for e in events if e.source_id == sid and e.event_date is not None),
            default=None,
        )
        last_str = last_date.isoformat() if last_date else "—"
        suffix = f"  error: {entry['error']}" if entry["error"] else ""
        out.append(f"  {sid:<28} status={status:<22} count={count:<5} last={last_str}{suffix}")
    out.append("")
    return out


def _section_events_per_month(events: list[EventRecord], num_months: int = 24) -> list[str]:
    if not events:
        return ["== Events per month per source (last 24 months) ==", "  (no events)", ""]

    today = date.today()
    months: list[tuple[int, int]] = []
    y, m = today.year, today.month
    for _ in range(num_months):
        months.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    months.reverse()

    sources = sorted({e.source_id for e in events})
    counts: dict[tuple[int, int], dict[str, int]] = {
        ym: dict.fromkeys(sources, 0) for ym in months
    }
    for e in events:
        if e.event_date is None:
            continue
        ym = (e.event_date.year, e.event_date.month)
        if ym in counts:
            counts[ym][e.source_id] = counts[ym].get(e.source_id, 0) + 1

    col_width = max(max(len(s) for s in sources), 8) + 2
    out = ["== Events per month per source (last 24 months) =="]
    header = f"{'month':<10}" + "".join(f"{s:>{col_width}}" for s in sources)
    out.append(header)
    for y, m in months:
        row = f"{y:04d}-{m:02d}    " + "".join(
            f"{counts[(y, m)][s]:>{col_width}}" for s in sources
        )
        out.append(row)
    out.append("")
    return out


def _section_count_breakdown(
    heading: str,
    events: list[EventRecord],
    key: Any,
    *,
    limit: int | None = None,
) -> list[str]:
    counter: Counter[str] = Counter()
    for e in events:
        value = key(e)
        if value:
            counter[str(value)] += 1
    if not counter:
        return []
    out = [f"== {heading} =="]
    for label, n in counter.most_common(limit):
        out.append(f"  {label:<40} {n}")
    out.append("")
    return out


def main(argv: list[str] | None = None) -> int:
    """Entry point for `python -m lawtracker.scout`."""
    parser = argparse.ArgumentParser(prog="lawtracker.scout", description="Run the data scout.")
    parser.add_argument("--source", help="Only run the named adapter (source_id)")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args(argv)

    report = run(PILOT_ADAPTERS, args.output_dir, source_filter=args.source)
    _print_report(report)
    return 0


def _print_report(report: dict[str, Any]) -> None:
    print(
        f"Scout complete: {report['events_collected']} events written to {report['output_dir']}"
    )
    for entry in report["poll_log"]:
        sid = entry["source_id"]
        status = entry["status"]
        count = entry["event_count"]
        suffix = f"  error: {entry['error']}" if entry["error"] else ""
        print(f"  {sid:<28} status={status:<22} count={count}{suffix}")


if __name__ == "__main__":
    raise SystemExit(main())
